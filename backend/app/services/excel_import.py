from __future__ import annotations

from io import BytesIO
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.config import HM_POS_PATTERNS, PO_UNIT_PATTERNS, SEED_FORM_HM, STATUS_PATTERNS
from app.models import DataHmRow, MasterUnit, MasterVendor, PoUnitRow, StatusRow


def find_sheet(sheetnames: list[str], patterns: tuple[str, ...]) -> str | None:
    upper = {s.upper(): s for s in sheetnames}
    for p in patterns:
        if p.upper() in upper:
            return upper[p.upper()]
    for s in sheetnames:
        su = s.upper()
        for p in patterns:
            if su.startswith(p.upper()):
                return s
    return None


def seed_master_from_form(db: Session, path: Path | None = None) -> int:
    path = path or SEED_FORM_HM
    if not path.exists():
        # fallback minimal vendors
        defaults = {
            "PT. JUS": [],
            "PT. SIE": [],
            "PT. WPK": [],
            "PT. WWPI": [],
            "PT. PIK": [],
        }
        count = 0
        for vendor, units in defaults.items():
            if not db.query(MasterVendor).filter_by(name=vendor).first():
                db.add(MasterVendor(name=vendor))
            for u in units:
                if not db.query(MasterUnit).filter_by(vendor=vendor, code_unit=u).first():
                    db.add(MasterUnit(vendor=vendor, code_unit=u))
                    count += 1
        db.commit()
        return count

    wb = load_workbook(path, read_only=True, data_only=True)
    if "Master Unit" not in wb.sheetnames:
        wb.close()
        return 0
    ws = wb["Master Unit"]
    # Row 2 = vendors in cols B..
    vendors: dict[int, str] = {}
    for col in range(2, (ws.max_column or 2) + 1):
        name = ws.cell(2, col).value
        if name:
            vendors[col] = str(name).strip()
            if not db.query(MasterVendor).filter_by(name=vendors[col]).first():
                db.add(MasterVendor(name=vendors[col]))

    count = 0
    for col, vendor in vendors.items():
        for row in range(3, (ws.max_row or 3) + 1):
            unit = ws.cell(row, col).value
            if not unit:
                continue
            code = str(unit).strip()
            exists = (
                db.query(MasterUnit)
                .filter_by(vendor=vendor, code_unit=code)
                .first()
            )
            if not exists:
                db.add(MasterUnit(vendor=vendor, code_unit=code))
                count += 1
    db.commit()
    wb.close()
    return count


def import_status_excel(db: Session, content: bytes, replace: bool = True) -> int:
    wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
    sheet = find_sheet(wb.sheetnames, STATUS_PATTERNS)
    if not sheet:
        wb.close()
        raise ValueError(f"Sheet STATUS tidak ditemukan. Sheets: {wb.sheetnames}")
    ws = wb[sheet]
    # Find header row containing Date / Code Unit
    header_row = None
    headers: dict[int, str] = {}
    for r in range(1, min(20, (ws.max_row or 1) + 1)):
        vals = {c: ws.cell(r, c).value for c in range(1, min(30, (ws.max_column or 1) + 1))}
        text = " ".join(str(v).upper() for v in vals.values() if v)
        if "CODE UNIT" in text and ("DATE" in text or "TANGGAL" in text):
            header_row = r
            headers = {c: str(v).strip() for c, v in vals.items() if v}
            break
    if header_row is None:
        wb.close()
        raise ValueError("Header STATUS (Date / Code Unit) tidak ditemukan")

    col_map = {name.upper(): idx for idx, name in headers.items()}

    def col(*names: str) -> int | None:
        for n in names:
            if n.upper() in col_map:
                return col_map[n.upper()]
        for key, idx in col_map.items():
            for n in names:
                if n.upper() in key:
                    return idx
        return None

    c_date = col("Date", "DATE", "TANGGAL")
    c_unit = col("Code Unit", "CODE UNIT")
    c_cat = col("Category")
    c_item = col("Item Category", "ITEM CATEGORY")
    c_jam = col("Jam", "JAM")
    c_shift = col("Shift", "SHIFT")
    c_area = col("Working Area")
    c_remarks = col("Remarks / Keterangan", "Remarks", "Keterangan")
    c_lokasi = col("Lokasi", "LOKASI")
    c_awal = col("Awal")
    c_akhir = col("Akhir")
    c_equip = col("Equipment")

    if replace:
        db.query(StatusRow).delete()

    n = 0
    for r in range(header_row + 1, (ws.max_row or header_row) + 1):
        unit = ws.cell(r, c_unit).value if c_unit else None
        if unit is None or str(unit).strip() == "":
            continue
        dval = ws.cell(r, c_date).value if c_date else None
        d = pd.to_datetime(dval, errors="coerce")
        jam_raw = ws.cell(r, c_jam).value if c_jam else 0
        try:
            jam = float(jam_raw or 0)
        except (TypeError, ValueError):
            jam = 0.0
        db.add(StatusRow(
            date=None if pd.isna(d) else d.date(),
            code_unit=str(unit).strip(),
            category=str(ws.cell(r, c_cat).value or "").strip() if c_cat else None,
            item_category=str(ws.cell(r, c_item).value or "").strip() if c_item else None,
            awal=str(ws.cell(r, c_awal).value) if c_awal and ws.cell(r, c_awal).value else None,
            akhir=str(ws.cell(r, c_akhir).value) if c_akhir and ws.cell(r, c_akhir).value else None,
            jam=jam,
            working_area=str(ws.cell(r, c_area).value or "").strip() if c_area else None,
            remarks=str(ws.cell(r, c_remarks).value or "").strip() if c_remarks else None,
            shift=str(ws.cell(r, c_shift).value or "").strip() if c_shift else None,
            equipment=str(ws.cell(r, c_equip).value or "").strip() if c_equip else None,
            lokasi=str(ws.cell(r, c_lokasi).value or "").strip() if c_lokasi else None,
        ))
        n += 1
    db.commit()
    wb.close()
    return n


def import_po_excel(db: Session, content: bytes, replace: bool = True) -> int:
    wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
    sheet = find_sheet(wb.sheetnames, PO_UNIT_PATTERNS)
    if not sheet:
        wb.close()
        raise ValueError(f"Sheet PO Unit tidak ditemukan. Sheets: {wb.sheetnames}")
    ws = wb[sheet]
    header_row = None
    headers: dict[int, str] = {}
    for r in range(1, min(15, (ws.max_row or 1) + 1)):
        vals = {c: ws.cell(r, c).value for c in range(1, min(35, (ws.max_column or 1) + 1))}
        text = " ".join(str(v).upper() for v in vals.values() if v)
        if "CODE UNIT" in text or "LAST PO" in text:
            header_row = r
            headers = {c: str(v).strip() for c, v in vals.items() if v}
            break
    if header_row is None:
        wb.close()
        raise ValueError("Header PO Unit tidak ditemukan")

    col_map = {name.upper(): idx for idx, name in headers.items()}

    def col(*names: str) -> int | None:
        for n in names:
            if n.upper() in col_map:
                return col_map[n.upper()]
        return None

    c_vendor = col("VENDOR")
    c_po = col("LAST PO", "PO NUMBER", "FIRST PO")
    c_unit = col("UNIT", "TYPE", "EQUIPMENT")
    c_year = col("TAHUN UNIT", "YEAR")
    c_code = col("CODE UNIT", "CODE UNIT MCR")
    c_periode = col("PERIODE PO", "PERIODE")
    c_no = col("NO UNIT")

    if replace:
        db.query(PoUnitRow).delete()

    n = 0
    for r in range(header_row + 1, (ws.max_row or header_row) + 1):
        code = ws.cell(r, c_code).value if c_code else None
        if not code:
            continue
        db.add(PoUnitRow(
            vendor=str(ws.cell(r, c_vendor).value or "").strip() if c_vendor else None,
            po_number=str(ws.cell(r, c_po).value or "").strip() if c_po else None,
            equipment=str(ws.cell(r, c_unit).value or "").strip() if c_unit else None,
            year=str(ws.cell(r, c_year).value or "").strip() if c_year else None,
            code_unit_mcr=str(code).strip(),
            periode_str=str(ws.cell(r, c_periode).value or "").strip() if c_periode else None,
            no_unit=str(ws.cell(r, c_no).value or "").strip() if c_no else None,
        ))
        n += 1
    db.commit()
    wb.close()
    return n


def _to_time(v):
    from datetime import datetime, time

    if v is None or (isinstance(v, float) and pd.isna(v)) or v == "":
        return None
    if isinstance(v, time):
        return v
    if isinstance(v, datetime):
        return v.time()
    if isinstance(v, pd.Timestamp):
        return v.to_pydatetime().time()
    if isinstance(v, (int, float)):
        total = int(round(float(v) * 24 * 3600)) % (24 * 3600)
        h, rem = divmod(total, 3600)
        m, s = divmod(rem, 60)
        return time(h, m, s)
    s = str(v).strip()
    for fmt in ("%H:%M:%S", "%H:%M"):
        try:
            return datetime.strptime(s, fmt).time()
        except ValueError:
            continue
    return None


def _to_float(v, default: float = 0.0) -> float:
    if v is None or (isinstance(v, float) and pd.isna(v)) or v == "":
        return default
    try:
        return float(v)
    except (TypeError, ValueError):
        return default


def import_data_hm_excel(db: Session, content: bytes, replace: bool = True) -> int:
    """Import sheet 'DATA HM' dari workbook BA Monitoring / form HM."""
    xl = pd.ExcelFile(BytesIO(content), engine="openpyxl")
    sheet = find_sheet(list(xl.sheet_names), HM_POS_PATTERNS)
    if not sheet:
        raise ValueError(f"Sheet DATA HM tidak ditemukan. Sheets: {list(xl.sheet_names)}")

    df = pd.read_excel(xl, sheet_name=sheet, header=1)
    df.columns = [str(c).strip() for c in df.columns]
    if "CODE UNIT" not in df.columns or "DATE" not in df.columns:
        raise ValueError(f"Kolom CODE UNIT / DATE tidak ada di sheet {sheet}")

    df = df.dropna(subset=["CODE UNIT", "DATE"])
    df["DATE"] = pd.to_datetime(df["DATE"], errors="coerce")
    df = df.dropna(subset=["DATE"])
    df["SHIFT"] = df.get("SHIFT", "Shift 1").astype(str).str.strip()
    df["VENDOR"] = df.get("VENDOR", "").astype(str).str.strip()
    df["CODE UNIT"] = df["CODE UNIT"].astype(str).str.strip()
    df = df.drop_duplicates(subset=["DATE", "SHIFT", "CODE UNIT"], keep="first")

    if replace:
        db.query(DataHmRow).delete()
        db.commit()

    n = 0
    for _, r in df.iterrows():
        located = r.get("LOCATED")
        job = r.get("JOB DESCRIPTION")
        op = r.get("OPERATORE NAME")
        ket = r.get("Keterangan")
        exp = r.get("EXP. DIFFERENCE")
        db.add(DataHmRow(
            date=r["DATE"].date(),
            shift=r["SHIFT"] or "Shift 1",
            vendor=r["VENDOR"] or "",
            code_unit=r["CODE UNIT"],
            hm_start=_to_float(r.get("HM START")),
            hm_stop=_to_float(r.get("HM STOP")),
            hours_start=_to_time(r.get("HOURS START")),
            hours_stop=_to_time(r.get("HOURS STOP")),
            jam_bd=_to_float(r.get("JAM BD")),
            jam_standby=_to_float(r.get("JAM STANDBY")),
            ritase=_to_float(r.get("RITASE")),
            fuel=_to_float(r.get("Fuel", r.get("FUEL"))),
            hm_pengisian=_to_float(r.get("HM Pengisian ", r.get("HM Pengisian"))),
            located=(
                str(located).strip()
                if pd.notna(located) and str(located).strip() not in ("", "0")
                else None
            ),
            job_description=str(job).strip() if pd.notna(job) else None,
            operator_name=(
                str(op).strip()
                if pd.notna(op) and str(op).strip() not in ("", "0")
                else None
            ),
            keterangan=(
                str(ket).strip()
                if pd.notna(ket) and str(ket).strip() not in ("", "0")
                else None
            ),
            exp_difference=(
                str(exp).strip()
                if pd.notna(exp) and str(exp).strip()
                else None
            ),
        ))
        n += 1
        if n % 500 == 0:
            db.commit()
    db.commit()
    return n