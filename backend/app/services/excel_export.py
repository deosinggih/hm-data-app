from __future__ import annotations

from io import BytesIO

import pandas as pd
from openpyxl import Workbook
from sqlalchemy.orm import Session

from app.models import DataHmRow, MasterUnit, StatusRow


def _num2(v):
    """Round numeric values to whole numbers for Excel export."""
    if v is None or v == "":
        return v
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        return int(round(float(v)))
    return v


FORM_HM_HEADERS = [
    "DATE", "SHIFT", "VENDOR", "CODE UNIT LAPANGAN", "CODE UNIT", "CN",
    "HM START", "HM STOP", "AMOUNT (HM)", "HOURS START", "HOURS STOP",
    "AMOUNT (EW)", "JAM BD", "JAM STANDBY", "RITASE", "Fuel", "HM Pengisian ",
    "LOCATED", "JOB DESCRIPTION", "OPERATORE NAME", "Keterangan",
    "HM DIFFERENCE", "EXP. DIFFERENCE", "INFORMATION", "HM TODAY",
    "EWH", "STB", "BD", "HM Pemotongan", "Remaks",
]

INPUT_MARKS = {
    "DATE", "SHIFT", "VENDOR", "CODE UNIT", "HM START", "HM STOP",
    "HOURS START", "HOURS STOP", "JAM BD", "JAM STANDBY", "RITASE",
    "Fuel", "HM Pengisian ", "LOCATED", "JOB DESCRIPTION",
    "OPERATORE NAME", "Keterangan", "EXP. DIFFERENCE",
}


def export_form_hm_xlsx(db: Session) -> bytes:
    """Export format form HM.xlsx: Master Unit + DATA HM sorted by DATE → SHIFT → CODE UNIT."""
    wb = Workbook()

    # Master Unit
    ws_m = wb.active
    ws_m.title = "Master Unit"
    units = db.query(MasterUnit).order_by(MasterUnit.vendor, MasterUnit.code_unit).all()
    by_vendor: dict[str, list[str]] = {}
    for u in units:
        by_vendor.setdefault(u.vendor, []).append(u.code_unit)
    vendors = list(by_vendor.keys())
    for i, v in enumerate(vendors):
        ws_m.cell(2, i + 2, v)
        for j, code in enumerate(by_vendor[v]):
            ws_m.cell(3 + j, i + 2, code)

    # DATA HM - Clean format (no grouping headers)
    ws = wb.create_sheet("DATA HM")

    # Row 1: Headers (DATE, SHIFT, VENDOR, CODE UNIT, CN, HM START, HM STOP, AMOUNT (HM), ...)
    header_columns = ["DATE", "SHIFT", "VENDOR", "CODE UNIT", "CN", "HM START", "HM STOP", "AMOUNT (HM)"]
    for i, h in enumerate(header_columns):
        ws.cell(1, i + 1, h)

    # Row 2: Empty row with 0.0 in AMOUNT column
    ws.cell(2, 8, 0.0)

    # Get rows sorted by DATE → CODE UNIT → SHIFT → HM START (sequential HM start untuk multiple operators same shift)
    rows = db.query(DataHmRow).order_by(
        DataHmRow.date.asc(),
        DataHmRow.code_unit.asc(),
        DataHmRow.shift.asc(),
        DataHmRow.hm_start.asc(),
    ).all()

    # Add data rows starting from row 3
    for r_i, row in enumerate(rows, start=3):
        # Format DATE as "1-Jun" format
        date_str = row.date.strftime("%-d-%b") if row.date else ""

        values = [
            date_str,                          # DATE (formatted as "1-Jun")
            row.shift,                         # SHIFT
            row.vendor,                        # VENDOR
            row.code_unit,                     # CODE UNIT
            row.cn or "",                      # CN
            row.hm_start,                      # HM START
            row.hm_stop,                       # HM STOP
            row.amount_hm if row.amount_hm is not None else 0.0,  # AMOUNT (HM)
        ]
        for c_i, val in enumerate(values):
            cell = ws.cell(r_i, c_i + 1)
            if isinstance(val, (int, float)):
                cell.value = float(val)
            else:
                cell.value = val

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_pa_excel(recap: pd.DataFrame, vendor_summary: pd.DataFrame | None = None) -> bytes:
    """Export Summary PA dengan header yang sama seperti UI."""
    if recap is None or recap.empty:
        out = pd.DataFrame(columns=[
            "Vendor", "Code Unit", "Total HM Working", "Total Breakdown",
            "Total Standby", "Total Forcemaejure", "Persentage PA",
        ])
    else:
        out = pd.DataFrame({
            "Vendor": recap.get("VENDOR"),
            "Code Unit": recap.get("CODE UNIT"),
            "Total HM Working": recap.get("WORKING"),
            "Total Breakdown": recap.get("BD"),
            "Total Standby": recap.get("STBY"),
            "Total Forcemaejure": recap.get("FM"),
            "Persentage PA": recap.get("PA (%)"),
        })
        for col in ("Total HM Working", "Total Breakdown", "Total Standby", "Total Forcemaejure", "Persentage PA"):
            if col in out.columns:
                out[col] = pd.to_numeric(out[col], errors="coerce").round(0)
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="xlsxwriter") as writer:
        out.to_excel(writer, sheet_name="Summary PA Unit", index=False)
    return buf.getvalue()


def export_status_browse_xlsx(db: Session) -> bytes:
    rows = db.query(StatusRow).all()
    data = [{
        "Date": r.date,
        "Code Unit": r.code_unit,
        "Category": r.category,
        "Item Category": r.item_category,
        "Jam": r.jam,
        "Shift": r.shift,
        "Working Area": r.working_area,
        "Remarks": r.remarks,
        "Lokasi": r.lokasi,
    } for r in rows]
    buf = BytesIO()
    pd.DataFrame(data).to_excel(buf, sheet_name="STATUS", index=False, engine="openpyxl")
    return buf.getvalue()